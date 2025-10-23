#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
union.py
--------
Unifica exports de Sprinklr, YouScan y Tubular en un único .xlsx:
- Hojas por fuente (sprinklr, tubular, youscan)
- Hoja "combined_agg" agregada por (date, hora, source, sentiment)
- Formatos Excel forzados:
  * date: [$-es-CO]dd/mm/yy  (equiv. visual a "Fecha corta")
  * hora y hour_original: [$-es-CO]h:mm:ss AM/PM
"""

from __future__ import annotations

import argparse
import glob
import os
from pathlib import Path
from typing import Iterable, List, Optional, Sequence, Tuple, Dict

import pandas as pd

try:
    import pytz
except Exception:
    pytz = None  # se maneja abajo con chequeo

# xlsxwriter opcional (si no está, se usa openpyxl)
try:
    import xlsxwriter  # noqa: F401
    HAS_XLSXWRITER = True
except Exception:
    HAS_XLSXWRITER = False

# Diccionario ISO-2 -> Nombre de país (español)
COUNTRY_CODES = {
    "CO": "Colombia", "MX": "Mexico", "AR": "Argentina", "BR": "Brazil",
    "CL": "Chile", "PE": "Peru", "EC": "Ecuador", "VE": "Venezuela",
    "US": "USA", "CA": "Canada"
    # Agrega más según necesites
}

def _expand_country_code(series: pd.Series) -> pd.Series:
    """
    Convierte códigos ISO-2 a nombres completos.
    Ej: 'CO' -> 'Colombia', 'mx' -> 'México'
    Si no encuentra el código, devuelve el valor original.
    """
    return series.str.upper().map(lambda x: COUNTRY_CODES.get(x, x) if pd.notna(x) else x)
# -----------------------------
# Configuración y utilidades
# -----------------------------

CANON_COLUMNS: List[str] = [
    "date",            # fecha pura (python date) -> Excel dd/mm/yy
    "hora",            # hora FLOOR(h) como fracción de día -> Excel h:mm:ss AM/PM
    "hour_original",   # hora original como fracción de día -> Excel h:mm:ss AM/PM
    "author",
    "message",
    "link",
    "source",
    "sentiment",
    "country",
    "engagement",
    "reach",
    "views",
    "mentions",
    "original_file",
]

# sinónimos por campo y por fuente
SYN_SPRINKLR: Dict[str, Sequence[str]] = {
    "author": ["From User","User Name","Author","Author Name","User"],
    "message": ["Conversation Stream","Message","Text","Content","Message Text","Post Message"],
    "link": ["Post Url","URL","Link","Permalink"],
    "created": ["Created Time.1","Created At","Fecha","Timestamp","Published Date","Published_Date","Created Time"],
    "source": ["snTypeColumn","Source","Source Type","Network"],
    "sentiment": ["Sentiment","Recalibrated Sentiment","Post Sentiment","Message Sentiment"],
    "country": ["Country","Author Country","Profile Country"],
    "reach": ["Reach (SUM)","Reach","Total Reach"],
    "engagement": ["Earned Engagements (Recalibrated) (SUM)","Engagements","Total Engagement","Interactions"],
    "mentions": ["Mentions (SUM)","Mentions","Count"],
}

SYN_YOUSCAN: Dict[str, Sequence[str]] = {
    "author": ["Author","Author Name","Nickname","User"],
    "message": ["Text","Message","Content","Text snippet"],
    "link": ["Link","URL","Permalink"],
    "date": ["Date"],   # dd.mm.yyyy
    "time": ["Time"],   # HH:MM 24h
    "source": ["snTypeColumn","Source","Source Type","Network"],
    "sentiment": ["Sentiment"],
    "country": ["Country","Location"],
    "engagement": ["Engagement","Interactions","Total engagement"],
    "views": ["Views","View count"],
    "mentions": ["Mentions","Count"],
}

SYN_TUBULAR: Dict[str, Sequence[str]] = {
    "author": ["Channel Name","Creator","Owner","Author"],
    "message": ["Title","Video_Title","Description"],
    "link": ["Video Link","URL","Link","Permalink","Video_URL"],
    "created": ["Published Date","Published_Date","Created Time","Timestamp","Fecha"],
    "country": ["Country","Owner Country","Channel Country","Creator_Country","Creator Country"],
    "views": ["Views","View Count","Total Views"],
    "source": ["snTypeColumn","Source","Source Type","Network","Platform"],
    "engagement": ["Engagement","Total_Engagements","Interactions","Reactions"],
    "mentions": ["Mentions","Count"],
}

# -----------------------------
# Helpers
# -----------------------------

def _detect_is_excel(path: str) -> bool:
    ext = Path(path).suffix.lower()
    return ext in (".xlsx", ".xls", ".xlsm", ".xlsb")

def read_any(path: str, skiprows: int = 0, header: int = 0, dtype: Optional[Dict] = None) -> pd.DataFrame:
    if _detect_is_excel(path):
        return pd.read_excel(path, skiprows=skiprows, header=header, dtype=dtype)
    for enc in ("utf-8-sig", "utf-8", "latin-1", "utf-16", "cp1252"):
        try:
            return pd.read_csv(path, skiprows=skiprows, header=header, dtype=dtype, encoding=enc)
        except Exception:
            continue
    return pd.read_csv(path, skiprows=skiprows, header=header, dtype=dtype, engine="python")

def _normalize_cols(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy()
    df2.columns = [str(c).strip() for c in df2.columns]
    return df2

def _first_match(cols: Iterable[str], candidates: Sequence[str]) -> Optional[str]:
    norm = {c.strip().lower(): c for c in cols}
    for cand in candidates:
        key = cand.strip().lower()
        if key in norm:
            return norm[key]
    return None

def _ensure_tz(dt: pd.Series, tz_from: Optional[str], tz_to: Optional[str]) -> Tuple[pd.Series, pd.Series, pd.Series]:
    """
    Devuelve (date_str_24h, time_str_24h, dt_converted) con TZ aplicada si corresponde.
    * date_str_24h: YYYY-MM-DD (utilitario)
    * time_str_24h: HH:MM (utilitario)
    * dt_converted: datetime64[ns, tz] para cálculos/formatos
    """
    if tz_from and tz_to and pytz is not None:
        try:
            src = pytz.timezone(tz_from)
            dst = pytz.timezone(tz_to)
            if getattr(dt.dt, "tz", None) is None:
                dt_localized = dt.dt.tz_localize(src, nonexistent='NaT', ambiguous='NaT')
            else:
                dt_localized = dt
            dt_converted = dt_localized.dt.tz_convert(dst)
        except Exception:
            dt_converted = dt
    else:
        dt_converted = dt
    date_str = dt_converted.dt.strftime("%Y-%m-%d")
    time_str = dt_converted.dt.strftime("%H:%M")
    return date_str, time_str, dt_converted

def _coerce_datetime(series: pd.Series, dayfirst: bool = False) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", dayfirst=dayfirst, infer_datetime_format=True)

def _excel_time_fraction(dt_series: pd.Series) -> pd.Series:
    """
    Convierte datetime (naive) a fracción de día Excel (float en [0,1)),
    preservando sólo hora/min/seg. NaT -> NaN.
    """
    hours = dt_series.dt.hour.fillna(0)
    mins  = dt_series.dt.minute.fillna(0)
    secs  = dt_series.dt.second.fillna(0)
    return (hours*3600 + mins*60 + secs) / 86400.0

def _clean_source(series: pd.Series, *, youscan: bool = False, default_value: str = "") -> pd.Series:
    """
    Normaliza 'source':
    - Siempre: minúsculas y trim.
    - YouScan: elimina 'http(s)://', 'www.', TLDs (.com/.co/.org/.net/.io/.app/.tv/.info) y paths.
               Ej: 'https://www.instagram.com/xyz' -> 'instagram'
    """
    s = series.fillna(default_value).astype(str).str.strip().str.lower()
    if youscan:
        s = (
            s.str.replace(r"^https?://", "", regex=True)
             .str.replace(r"^www\.", "", regex=True)
             .str.replace(r"/.*$", "", regex=True)                    # corta en el primer '/'
             .str.replace(r"\.(com|co|org|net|io|app|tv|info).*$", "", regex=True)
             .str.replace(r"\.$", "", regex=True)
             .str.strip()
        )
    return s

# -----------------------------
# Procesamiento por FUENTE
# -----------------------------

def process_sprinklr(paths: Sequence[str], skiprows: int = 0, header: int = 0,
                     tz_from: Optional[str] = None, tz_to: Optional[str] = None) -> pd.DataFrame:
    rows = []
    for p in paths:
        df = _normalize_cols(read_any(p, skiprows=skiprows, header=header))
        cols = list(df.columns)

        c_author  = _first_match(cols, SYN_SPRINKLR["author"])
        c_msg     = _first_match(cols, SYN_SPRINKLR["message"])
        c_link    = _first_match(cols, SYN_SPRINKLR["link"])
        c_created = _first_match(cols, SYN_SPRINKLR["created"])
        c_source  = _first_match(cols, SYN_SPRINKLR["source"])
        c_sent    = _first_match(cols, SYN_SPRINKLR["sentiment"])
        c_country = _first_match(cols, SYN_SPRINKLR["country"])
        c_reach   = _first_match(cols, SYN_SPRINKLR["reach"])
        c_eng     = _first_match(cols, SYN_SPRINKLR["engagement"])
        c_mentions= _first_match(cols, SYN_SPRINKLR["mentions"])

        tmp = pd.DataFrame()
        tmp["original_file"] = Path(p).name
        tmp["author"]   = df.get(c_author,  pd.Series([None]*len(df)))
        tmp["message"]  = df.get(c_msg,     pd.Series([None]*len(df)))
        tmp["link"]     = df.get(c_link,    pd.Series([None]*len(df)))

        created_raw = df.get(c_created, pd.Series([None]*len(df)))
        created_dt  = _coerce_datetime(created_raw, dayfirst=False)
        _d24, _t24, dt_conv = _ensure_tz(created_dt, tz_from, tz_to)

        tmp["date"]          = dt_conv.dt.tz_localize(None).dt.normalize()
        tmp["hora"]          = _excel_time_fraction(dt_conv.dt.floor("h").dt.tz_localize(None))
        tmp["hour_original"] = _excel_time_fraction(dt_conv.dt.tz_localize(None))

        src_series = df.get(c_source, pd.Series(["sprinklr"]*len(df)))
        tmp["source"] = _clean_source(src_series, youscan=False, default_value="sprinklr")

        tmp["sentiment"]  = df.get(c_sent,    pd.Series([None]*len(df)))
        tmp["country"]    = df.get(c_country, pd.Series([None]*len(df)))
        tmp["engagement"] = df.get(c_eng,     pd.Series([None]*len(df)))
        tmp["reach"]      = df.get(c_reach,   pd.Series([None]*len(df)))
        tmp["views"]      = None
        tmp["mentions"]   = df.get(c_mentions,pd.Series([1]*len(df)))
        rows.append(tmp)

    out = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame(columns=CANON_COLUMNS)
    return out.reindex(columns=CANON_COLUMNS)

def process_youscan(paths: Sequence[str], skiprows: int = 0, header: int = 0,
                    tz_from: Optional[str] = None, tz_to: Optional[str] = None) -> pd.DataFrame:
    rows = []
    for p in paths:
        df = _normalize_cols(read_any(p, skiprows=skiprows, header=header))
        cols = list(df.columns)

        c_author  = _first_match(cols, SYN_YOUSCAN["author"])
        c_msg     = _first_match(cols, SYN_YOUSCAN["message"])
        c_link    = _first_match(cols, SYN_YOUSCAN["link"])
        c_date    = _first_match(cols, SYN_YOUSCAN["date"])
        c_time    = _first_match(cols, SYN_YOUSCAN["time"])
        c_source  = _first_match(cols, SYN_YOUSCAN["source"])
        c_sent    = _first_match(cols, SYN_YOUSCAN["sentiment"])
        c_country = _first_match(cols, SYN_YOUSCAN["country"])
        c_eng     = _first_match(cols, SYN_YOUSCAN["engagement"])
        c_views   = _first_match(cols, SYN_YOUSCAN["views"])
        c_mentions= _first_match(cols, SYN_YOUSCAN["mentions"])

        tmp = pd.DataFrame()
        tmp["original_file"] = Path(p).name
        tmp["author"]   = df.get(c_author,  pd.Series([None]*len(df)))
        tmp["message"]  = df.get(c_msg,     pd.Series([None]*len(df)))
        tmp["link"]     = df.get(c_link,    pd.Series([None]*len(df)))

        # Date dd.mm.yyyy + Time HH:MM
        date_raw = df.get(c_date, pd.Series([None]*len(df)))
        time_raw = df.get(c_time, pd.Series([None]*len(df)))
        dt_combined = pd.to_datetime(
            date_raw.astype(str).str.strip() + " " + time_raw.astype(str).str.strip(),
            errors="coerce", dayfirst=True, infer_datetime_format=True
        )
        _d24, _t24, dt_conv = _ensure_tz(dt_combined, tz_from, tz_to)

        tmp["date"]          = dt_conv.dt.tz_localize(None).dt.normalize()
        tmp["hora"]          = _excel_time_fraction(dt_conv.dt.floor("h").dt.tz_localize(None))
        tmp["hour_original"] = _excel_time_fraction(dt_conv.dt.tz_localize(None))

        src_series = df.get(c_source, pd.Series(["youscan"]*len(df)))
        tmp["source"] = _clean_source(src_series, youscan=True, default_value="youscan")

        tmp["sentiment"]  = df.get(c_sent,    pd.Series([None]*len(df)))
        tmp["country"]    = df.get(c_country, pd.Series([None]*len(df)))
        tmp["engagement"] = df.get(c_eng,     pd.Series([None]*len(df)))
        tmp["reach"]      = None
        tmp["views"]      = df.get(c_views,   pd.Series([None]*len(df)))
        tmp["mentions"]   = df.get(c_mentions,pd.Series([1]*len(df)))
        rows.append(tmp)

    out = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame(columns=CANON_COLUMNS)
    return out.reindex(columns=CANON_COLUMNS)

def process_tubular(paths: Sequence[str], skiprows: int = 0, header: int = 0,
                    tz_from: Optional[str] = None, tz_to: Optional[str] = None) -> pd.DataFrame:
    rows = []
    for p in paths:
        df = _normalize_cols(read_any(p, skiprows=skiprows, header=header))
        cols = list(df.columns)

        c_author  = _first_match(cols, SYN_TUBULAR["author"])
        c_msg     = _first_match(cols, SYN_TUBULAR["message"])
        c_link    = _first_match(cols, SYN_TUBULAR["link"])
        c_created = _first_match(cols, SYN_TUBULAR["created"])
        c_source  = _first_match(cols, SYN_TUBULAR["source"])
        c_country = _first_match(cols, SYN_TUBULAR["country"])
        c_views   = _first_match(cols, SYN_TUBULAR["views"])
        c_eng     = _first_match(cols, SYN_TUBULAR["engagement"])
        c_mentions= _first_match(cols, SYN_TUBULAR["mentions"])

        tmp = pd.DataFrame()
        tmp["original_file"] = Path(p).name
        tmp["author"]   = df.get(c_author,  pd.Series([None]*len(df)))
        tmp["message"]  = df.get(c_msg,     pd.Series([None]*len(df)))
        tmp["link"]     = df.get(c_link,    pd.Series([None]*len(df)))

        created_raw = df.get(c_created, pd.Series([None]*len(df)))
        created_dt  = _coerce_datetime(created_raw, dayfirst=False)
        _d24, _t24, dt_conv = _ensure_tz(created_dt, tz_from, tz_to)

        tmp["date"]          = dt_conv.dt.tz_localize(None).dt.normalize()
        tmp["hora"]          = _excel_time_fraction(dt_conv.dt.floor("h").dt.tz_localize(None))
        tmp["hour_original"] = _excel_time_fraction(dt_conv.dt.tz_localize(None))

        src_series = df.get(c_source, pd.Series(["tubular"]*len(df)))
        tmp["source"] = _clean_source(src_series, youscan=False, default_value="tubular")

        # Country en Tubular suele venir ISO-2; normalizo a mayúsculas
        country_series = df.get(c_country, pd.Series([None]*len(df)))
        tmp["country"] = _expand_country_code(country_series.astype(str).str.strip()) #country_series.astype(str).str.strip().str.upper()

        tmp["sentiment"]  = None
        tmp["engagement"] = df.get(c_eng,     pd.Series([None]*len(df)))
        tmp["reach"]      = None
        tmp["views"]      = df.get(c_views,   pd.Series([None]*len(df)))
        tmp["mentions"]   = df.get(c_mentions,pd.Series([1]*len(df)))
        rows.append(tmp)

    out = pd.concat(rows, ignore_index=True) if rows else pd.DataFrame(columns=CANON_COLUMNS)
    return out.reindex(columns=CANON_COLUMNS)

# -----------------------------
# Motor principal
# -----------------------------

def _expand_globs(patterns: Sequence[str]) -> List[str]:
    out: List[str] = []
    for pat in patterns:
        out.extend(sorted(glob.glob(pat)))
    seen = set()
    uniq = []
    for p in out:
        if p not in seen:
            uniq.append(p); seen.add(p)
    return uniq

def etl_unify(sprinklr_files: Sequence[str] = (), tubular_files: Sequence[str] = (),
              youscan_files: Sequence[str] = (), out_xlsx: str = "etl_unificado.xlsx",
              skiprows_sprinklr: int = 0, skiprows_tubular: int = 0, skiprows_youscan: int = 0,
              header: int = 0, tz_from: Optional[str] = None, tz_to: Optional[str] = None) -> str:

    spr_paths = _expand_globs(sprinklr_files)
    tub_paths = _expand_globs(tubular_files)
    ysc_paths = _expand_globs(youscan_files)

    sheets = {}
    if spr_paths:
        sheets["sprinklr"] = process_sprinklr(spr_paths, skiprows=skiprows_sprinklr, header=header,
                                              tz_from=tz_from, tz_to=tz_to)
    if tub_paths:
        sheets["tubular"] = process_tubular(tub_paths, skiprows=skiprows_tubular, header=header,
                                            tz_from=tz_from, tz_to=tz_to)
    if ysc_paths:
        sheets["youscan"] = process_youscan(ysc_paths, skiprows=skiprows_youscan, header=header,
                                            tz_from=tz_from, tz_to=tz_to)

    if not sheets:
        raise ValueError("No se encontraron archivos de ninguna fuente (Sprinklr/Tubular/YouScan).")

    # Combine (solo para agregar) y orden base
    combined = pd.concat([sheets[k] for k in sheets], ignore_index=True).reindex(columns=CANON_COLUMNS)
    combined = combined.sort_values(["date", "hora"]).reset_index(drop=True)

    # ---- Agregación requerida: (date, hora, source, sentiment) ----
    agg_cols = ["mentions", "reach", "engagement", "views"]
    group_cols = ["date", "hora", "source", "sentiment", "country"]
    combined_agg = (
        combined.groupby(group_cols, dropna=False, as_index=False)[agg_cols]
        .sum(min_count=1)
        .sort_values(["date", "hora", "source", "country", "sentiment"])
        .reset_index(drop=True)
    )

    # -------- Guardado con formatos (xlsxwriter si está; si no, openpyxl) --------
    out_xlsx = str(out_xlsx)
    out_dir = os.path.dirname(out_xlsx)
    if out_dir and not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)

    DATE_FMT = "[$-es-CO]dd/mm/yy"
    TIME_FMT = "[$-es-CO]h:mm:ss AM/PM"

    if HAS_XLSXWRITER:
        with pd.ExcelWriter(out_xlsx, engine="xlsxwriter") as writer:
            wb = writer.book
            fmt_date = wb.add_format({"num_format": DATE_FMT})
            fmt_time = wb.add_format({"num_format": TIME_FMT})

            def _write_sheet(name: str, df: pd.DataFrame):
                df_sorted = df.sort_values(["date", "hora"]).reset_index(drop=True)
                df_sorted.to_excel(writer, sheet_name=name[:31], index=False)
                ws = writer.sheets[name[:31]]
                if "date" in df_sorted.columns:
                    c = df_sorted.columns.get_loc("date")
                    ws.set_column(c, c, 10, fmt_date)
                if "hora" in df_sorted.columns:
                    c = df_sorted.columns.get_loc("hora")
                    ws.set_column(c, c, 14, fmt_time)
                if "hour_original" in df_sorted.columns:
                    c = df_sorted.columns.get_loc("hour_original")
                    ws.set_column(c, c, 14, fmt_time)

            # Hojas por fuente:
            for name, df in sheets.items():
                _write_sheet(name, df)
            # Solo hoja agregada:
            _write_sheet("Total", combined_agg)

    else:
        from openpyxl import load_workbook
        with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
            def _write_sheet(name: str, df: pd.DataFrame):
                df_sorted = df.sort_values(["date", "hora"]).reset_index(drop=True)
                df_sorted.to_excel(writer, sheet_name=name[:31], index=False)
            for name, df in sheets.items():
                _write_sheet(name, df)
            _write_sheet("Total", combined_agg)

        wb = load_workbook(out_xlsx)

        def _format_ws(ws, df):
            if "date" in df.columns:
                col_idx = df.columns.get_loc("date") + 1
                for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                    for cell in col:
                        cell.number_format = DATE_FMT
            if "hora" in df.columns:
                col_idx = df.columns.get_loc("hora") + 1
                for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                    for cell in col:
                        cell.number_format = TIME_FMT
            if "hour_original" in df.columns:
                col_idx = df.columns.get_loc("hour_original") + 1
                for col in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2, max_row=ws.max_row):
                    for cell in col:
                        cell.number_format = TIME_FMT

        for name, df in list(sheets.items()) + [("combined_agg", Agregado)]:
            ws = wb[name[:31]] if name != "combined_agg" else wb["Agregado"]
            _format_ws(ws, df)
        wb.save(out_xlsx)

    return out_xlsx

# -----------------------------
# CLI
# -----------------------------

def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    p = argparse.ArgumentParser(
        description="Unifica Sprinklr, YouScan y Tubular en un .xlsx",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter
    )
    p.add_argument("--sprinklr", nargs="*", default=(), help="Rutas o patrones (glob) para Sprinklr, p.ej. '/content/data/spr_*.xlsx'")
    p.add_argument("--tubular", nargs="*", default=(), help="Rutas o patrones (glob) para Tubular, p.ej. '/content/data/tub_*.csv'")
    p.add_argument("--youscan", nargs="*", default=(), help="Rutas o patrones (glob) para YouScan, p.ej. '/content/data/ys_*.xlsx'")
    p.add_argument("--out", default="etl_unificado.xlsx", help="Ruta de salida .xlsx")
    p.add_argument("--skiprows_sprinklr", type=int, default=0, help="Filas a saltar al inicio (Sprinklr)")
    p.add_argument("--skiprows_tubular", type=int, default=0, help="Filas a saltar al inicio (Tubular)")
    p.add_argument("--skiprows_youscan", type=int, default=0, help="Filas a saltar al inicio (YouScan)")
    p.add_argument("--header", type=int, default=0, help="Fila de encabezado (misma para las 3 fuentes)")
    p.add_argument("--tz_from", default=None, help="Timezone de origen, p.ej. 'UTC'")
    p.add_argument("--tz_to", default=None, help="Timezone destino, p.ej. 'America/Bogota'")
    return p.parse_args(argv)

def main():
    args = parse_args()
    if (args.tz_from or args.tz_to) and pytz is None:
        raise RuntimeError("Se requiere 'pytz' para conversión de zonas horarias. Instala con: pip install pytz")

    result = etl_unify(
        sprinklr_files=args.sprinklr,
        tubular_files=args.tubular,
        youscan_files=args.youscan,
        out_xlsx=args.out,
        skiprows_sprinklr=args.skiprows_sprinklr,
        skiprows_tubular=args.skiprows_tubular,
        skiprows_youscan=args.skiprows_youscan,
        header=args.header,
        tz_from=args.tz_from,
        tz_to=args.tz_to,
    )
    print(f"\n✓ Archivo generado: {result}")

if __name__ == "__main__":
    main()
